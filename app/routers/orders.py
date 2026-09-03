from fastapi import APIRouter, Form, Depends, HTTPException
from fastapi.responses import StreamingResponse
from ..database import get_db
from ..utils import get_current_user
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from urllib.parse import quote


router = APIRouter()

def get_time():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")



# ========== 新增：导出 Excel ==========
@router.get("/orders/export")
def export_excel(user_id: int = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, type, amount, remark, created_at 
        FROM orders 
        WHERE user_id = %s 
        ORDER BY id DESC
    """, (user_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "账单记录"

    # 设置表头（加粗，灰色背景，居中）
    headers = ["ID", "类型", "金额", "备注", "时间"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["id"])
        ws.cell(row=row_idx, column=2, value="收入" if row["type"] == "income" else "支出")
        ws.cell(row=row_idx, column=3, value=float(row["amount"]))
        ws.cell(row=row_idx, column=4, value=row["remark"] or "")
        ws.cell(row=row_idx, column=5, value=row["created_at"].strftime("%Y-%m-%d %H:%M:%S") if row["created_at"] else "")

        # 金额列设置为数字格式
        ws.cell(row=row_idx, column=3).number_format = '#,##0.00'

        # 类型列居中
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")

    # 自动调整列宽
    #columns是 openpyxl 专门提供的一个列生成器。它会从 A 列开始，一列一列地把数据装进来。
    #为什么取 col[0]？因为一列里至少有一个格子（表头），取第一个格子的列字母就代表了这一整列的编号。这比遍历所有单元格再读字母高效得多。
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 6, 255)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存到内存字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回文件流
    filename = f"账单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)  # 对中文进行 URL 编码

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=export.xlsx; filename*=UTF-8''{encoded_filename}"
        }
    )


#刷新列表
@router.get("/orders")
def list_orders(user_id: int = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM orders WHERE user_id = %s ORDER BY id DESC", (user_id,))
    data = cur.fetchall()
    cur.close()
    conn.close()
    return {"code": 200, "data": data}


# 添加订单
@router.post("/orders")
def add_order(
    type: str = Form(...),
    amount: float = Form(...),
    remark: str = Form(""),
    user_id: int = Depends(get_current_user)
):
    if amount <= 0:
        raise HTTPException(status_code=400, detail="金额必须大于0")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO orders (type, amount, remark, user_id) VALUES (%s, %s, %s, %s)",
        (type, amount, remark, user_id)
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"code": 200, "msg": "添加成功", "time": get_time()}


#删除记录
@router.delete("/orders/{order_id}")
def delete_order(
    order_id: int,
    user_id: int = Depends(get_current_user)
):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM orders WHERE id = %s AND user_id = %s", (order_id, user_id))
    conn.commit()
    affected = cur.rowcount
    cur.close()
    conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="记录不存在或无权删除")
    return {"code": 200, "msg": "删除成功"}


# 修改订单
@router.put("/orders/{order_id}")
def update_order(
    order_id: int,
    type: str = Form(...),
    amount: float = Form(...),
    remark: str = Form(...),
    user_id: int = Depends(get_current_user)
):
    if amount <= 0:
        raise HTTPException(status_code=400, detail="金额必须大于0")
    
    conm = get_db()
    cur = conm.cursor()
    sql = "UPDATE orders SET type=%s, amount=%s, remark=%s WHERE id=%s AND user_id=%s"
    cur.execute(sql,(type, amount, remark, order_id, user_id))
    conm.commit()
    affected = cur.rowcount
    cur.close()
    conm.close()

    if affected == 0:
        raise HTTPException(status_code=404, detail="记录不存在或无权修改")
    return {"code": 200, "msg": "修改成功"}




#管理员界面
@router.get("/admin/orders")
def admin_list_orders(user_id: int = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    
    # 1. 强制校验：当前用户到底是不是管理员
    cur.execute("SELECT is_admin FROM users WHERE id = %s", (user_id,))
    result = cur.fetchone()
    if not result or result["is_admin"] != 1:
        raise HTTPException(status_code=403, detail="无权限访问管理后台")
    
    # 2. 查询所有用户的账单，并且关联用户名（管理员专用）
    cur.execute("""
        SELECT 
            orders.*,
            users.username
        FROM 
            orders 
        LEFT JOIN users ON orders.user_id = users.id 
        ORDER BY orders.id DESC
    """)
    data = cur.fetchall()
    cur.close()
    conn.close()
    
    return {"code": 200, "data": data}



@router.get("/admin/users")
def admin_list_users(user_id: int = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    
    # 验证是否为管理员
    cur.execute("SELECT is_admin FROM users WHERE id = %s", (user_id,))
    result = cur.fetchone()
    if not result or result["is_admin"] != 1:
        raise HTTPException(status_code=403, detail="无权限访问管理后台")
    
    # 获取所有用户（不返回密码哈希）
    cur.execute("SELECT id, username, is_admin, is_active, created_at FROM users ORDER BY id ASC")
    data = cur.fetchall()
    cur.close()
    conn.close()
    
    return {"code": 200, "data": data}


@router.put("/admin/users/{target_user_id}/toggle")
def toggle_user_status(
    target_user_id: int,
    user_id: int = Depends(get_current_user)
):
    conn = get_db()
    cur = conn.cursor()
    
    # 1. 验证当前用户是否为管理员
    cur.execute("SELECT is_admin FROM users WHERE id = %s", (user_id,))
    result = cur.fetchone()
    if not result or result["is_admin"] != 1:
        raise HTTPException(status_code=403, detail="无权限")
    
    # 2. 不能禁用自己（防止把管理员自己封了）
    if target_user_id == user_id:
        raise HTTPException(status_code=400, detail="不能禁用自己")
    
    # 3. 获取目标用户当前状态
    cur.execute("SELECT is_active FROM users WHERE id = %s", (target_user_id,))
    target = cur.fetchone()
    if not target:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 4. 切换状态（0→1 或 1→0）
    new_status = 1 if target["is_active"] == 0 else 0
    cur.execute(
        "UPDATE users SET is_active = %s WHERE id = %s",
        (new_status, target_user_id)
    )
    conn.commit()
    cur.close()
    conn.close()
    
    status_text = "启用" if new_status == 1 else "禁用"
    return {"code": 200, "msg": f"已{status_text}该用户"}



@router.get("/admin/orders/export")
def export_admin_orders(user_id: int = Depends(get_current_user)):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    from urllib.parse import quote

    conn = get_db()
    cur = conn.cursor()
    
    # 1. 验证管理员身份
    cur.execute("SELECT is_admin FROM users WHERE id = %s", (user_id,))
    result = cur.fetchone()
    if not result or result["is_admin"] != 1:
        raise HTTPException(status_code=403, detail="无权限")
    
    # 2. 关联查询所有用户的账单
    cur.execute("""
        SELECT orders.*, users.username
        FROM orders
        LEFT JOIN users ON orders.user_id = users.id
        ORDER BY orders.id DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # 3. 生成 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "所有用户账单"

    # 表头
    headers = ["ID", "类型", "金额", "备注", "时间", "用户名"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # ✅ 写入数据（这是你缺失的部分）
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["id"])
        ws.cell(row=row_idx, column=2, value="收入" if row["type"] == "income" else "支出")
        ws.cell(row=row_idx, column=3, value=float(row["amount"]))
        ws.cell(row=row_idx, column=4, value=row["remark"] or "")
        ws.cell(row=row_idx, column=5, value=row["created_at"].strftime("%Y-%m-%d %H:%M:%S") if row["created_at"] else "")
        ws.cell(row=row_idx, column=6, value=row["username"] or "已注销")
        
        # 金额列设置数字格式
        ws.cell(row=row_idx, column=3).number_format = '#,##0.00'

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 6, 255)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"所有账单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=export.xlsx; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.get("/admin/users/export")
def export_admin_users(user_id: int = Depends(get_current_user)):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    from urllib.parse import quote

    conn = get_db()
    cur = conn.cursor()
    
    # 1. 验证管理员身份
    cur.execute("SELECT is_admin FROM users WHERE id = %s", (user_id,))
    result = cur.fetchone()
    if not result or result["is_admin"] != 1:
        raise HTTPException(status_code=403, detail="无权限")
    
    # 2. 查询所有用户信息（不含密码）
    cur.execute("SELECT id, username, is_admin, is_active, created_at FROM users ORDER BY id ASC")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # 3. 生成 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户列表"

    # 表头
    headers = ["ID", "用户名", "角色", "状态", "注册时间"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # ✅ 写入数据（这是你缺失的部分）
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["id"])
        ws.cell(row=row_idx, column=2, value=row["username"])
        ws.cell(row=row_idx, column=3, value="管理员" if row["is_admin"] == 1 else "普通用户")
        ws.cell(row=row_idx, column=4, value="启用" if row["is_active"] == 1 else "禁用")
        ws.cell(row=row_idx, column=5, value=row["created_at"].strftime("%Y-%m-%d %H:%M:%S") if row["created_at"] else "")

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 6, 255)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"用户列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=export.xlsx; filename*=UTF-8''{encoded_filename}"
        }
    )